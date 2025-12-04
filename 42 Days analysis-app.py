import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile

st.title("42 Day Statistical Analysis - OMAC Developer")
st.subheader("Premium Interface")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file:
    start_col = st.text_input("Enter Start Column (e.g. Y)")
    end_col = st.text_input("Enter End Column (e.g. AA)")

    if st.button("Process File"):
        wb = load_workbook(uploaded_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Column indexes
            start_idx = ws[start_col + "1"].column
            end_idx = ws[end_col + "1"].column

            # Label column = one column before start column
            label_col = get_column_letter(start_idx - 1)

            # Insert 5 blank rows at the top
            ws.insert_rows(1, 5)

            # Recalculate and get last row
            last_row = ws.max_row

            # START formulas from row 7 (because row 6 is blank & row 7 is header)
            data_start_row = 7

            stats = [
                ("Mean", "AVERAGE"),
                ("Median", "MEDIAN"),
                ("Minimum", "MIN"),
                ("Maximum", "MAX"),
                ("Standard Deviation", "STDEV.S")
            ]

            write_row = 1

            for label, function in stats:
                # insert label e.g. Mean, Median...
                ws[f"{label_col}{write_row}"] = label

                # Insert formulas horizontally
                for col_idx in range(start_idx, end_idx + 1):
                    col_letter = get_column_letter(col_idx)
                    ws[f"{col_letter}{write_row}"] = f"={function}({col_letter}{data_start_row}:{col_letter}{last_row})"

                write_row += 1

        # Save File
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            st.success("File processed successfully!")
            st.download_button(
                "Download Updated Excel",
                data=tmp.read(),
                file_name="42_Day_Statistical_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
